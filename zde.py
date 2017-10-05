from pyzipcode import ZipCodeDatabase
from geopy.distance import vincenty
from openpyxl import load_workbook
from string import ascii_uppercase
from os import listdir, path
import sys


def col2num(col):
    """
    :param col: the string version of the column number you want. i.e A -> 1, B -> 2, C -> 3 etc
    :return: the index of the column (one-indexed)
    """

    num = 0
    for c in col:
        if c in ascii_uppercase:
            num = num * len(ascii_uppercase) + (ord(c) - ord(ascii_uppercase[0])) + 1
    return num


if __name__ == "__main__":

    files_in_dir = listdir(path.abspath(path.dirname(__file__)))

    xl_files_in_dir = []

    for file in files_in_dir:
        if file.endswith(".xlsx"):
            xl_files_in_dir.append(file)

    if len(xl_files_in_dir) <= 0:
        print("There aren't any .xlsx files in this directory! Exiting.")
        exit(0)

    print("Which file would you like to modify?")

    for file_number, file_name in enumerate(xl_files_in_dir):
        print("[" + str(file_number) + "] - ", file_name)

    file_number_string = input("File Number: ").rstrip()

    file_number = None

    try:
        file_number = int(file_number_string)
    except ValueError:
        print("[" + str(file_number_string) + "]", "Isn't a number. Exiting.")
        exit(0)

    selected_file_name = None

    try:
        selected_file_name = xl_files_in_dir[file_number]
    except IndexError:
        print("[" + str(file_number) + "]", "Doesn't correspond with a listed file number. Exiting.")

    print("You've selected [" + str(selected_file_name) + "]", "to edit.")

    wb = load_workbook(filename=selected_file_name)

    print("Which sheet would you like to modify?")

    for sheet_number, sheet_name in enumerate(wb.sheetnames):
        print("[" + str(sheet_number) + "] - ", sheet_name)

    sheet_number_string = input("Sheet Number: ").rstrip()

    sheet_number = None

    try:
        sheet_number = int(sheet_number_string)
    except ValueError:
        print("[" + str(sheet_number_string) + "]", "Isn't a number. Exiting.")
        exit(0)

    selected_sheet_name = None

    try:
        selected_sheet_name = wb.sheetnames[sheet_number]
    except IndexError:
        print("[" + str(file_number) + "]", "Doesn't correspond with a listed sheet number. Exiting.")

    print("You've selected [" + str(selected_sheet_name) + "]", "to edit.")

    read_col = input("Which column to read? (ie. A, B, AA): ").rstrip()
    write_col = input("Which column to write result? (ie. A, B, AA): ").rstrip()

    zipcode_database = ZipCodeDatabase()

    input_zip = input("Point A Zipcode? ").rstrip()

    a_zipcode = None

    try:
        a_zipcode = zipcode_database[input_zip]

    except IndexError as e:
        print("Couldn't find A zipcode:",
              str(input_zip),
              "in Database. Exiting.")
        exit(0)

    a_coords = (a_zipcode.latitude, a_zipcode.longitude)

    ws = wb[selected_sheet_name]

    modification_count = 0

    for row in ws.iter_rows(min_row=2):

        for cell in row:

                if cell.column == read_col:

                    if cell.value is not None:

                        lookup_zip = str(cell.value).zfill(5)  # there are 5 digits in a zipcode

                        try:
                            zipcode = zipcode_database[lookup_zip]

                            distance = vincenty(a_coords, (zipcode.latitude, zipcode.longitude)).miles

                            insertion_text = distance

                        except IndexError as e:
                            print("Couldn't find zipcode:",
                                  str(lookup_zip),
                                  "at cell: " + str(cell.column) + str(cell.row),
                                  "in Database")

                            insertion_text = "?"

                        target_column_index = col2num(write_col)
                        write_cell = ws.cell(row=cell.row, column=target_column_index)
                        write_cell.value = insertion_text
                        modification_count += 1
    try:
        wb.save(selected_file_name)
        print("Job Complete. " + str(modification_count) + " modifications made.")
    except PermissionError as e:
        print("The file [selected_file_name] is already open. Exiting.")
        exit(0)




